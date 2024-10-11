from openpyxl import load_workbook
import openpyxl
from tkinter import simpledialog, messagebox

class Ticket:
    filename = 'excel/ticket_info.xlsx' # ที่อยู่ของ file เก็บข้อมูลการซื่อตั๋ว

    def __init__(self, start_station, end_station, distance, price):
        self.Num_tic = 1
        self.start_station = start_station
        self.end_station = end_station
        self.distance = distance
        self.price = price

    # เพิ่มข้อมูลตั๋วลงตาราง
    def Save_To_Excel(self):

        try:
            # Try to load an existing workbook
            workbook = load_workbook(self.filename)
            sheet = workbook.active
        except FileNotFoundError:
            # If the file does not exist, create a new workbook and add headers
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            headers = ["Num_tic", "Start Station", "End Station", "Distance", "Price"]
            sheet.append(headers)
        # รันเลข row ต่อจากข้อมูลเก่า
        self.Num_tic = sheet.max_row

            # Add ticket data
        ticket_data = [self.Num_tic, self.start_station, self.end_station, self.distance, self.price]
        sheet.append(ticket_data)

        # Save the workbook
        workbook.save(self.filename)
        print(f"Ticket information saved to {self.filename}")

    # ลบข้อมูล clear data
    def delete_all_data(self, *args):
        password = simpledialog.askstring("Verify", "Enter password:", show='*')
        if password == "admin":
            try:
                # Load the workbook and select the active worksheet
                workbook = openpyxl.load_workbook(self.filename)
                sheet = workbook.active

                # Clear all rows except the header
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    for cell in row:
                        cell.value = None

                # Save the workbook
                workbook.save(self.filename)
                print(f"All data deleted from {self}")
            except FileNotFoundError:
                print(f"File not found: {self.filename}")
            except Exception as e:
                print(f"An error occurred: {e}")
        else:messagebox.showerror("Delete Data Failed", "Incorrect password")



    # เปิดใช้ฟังกืชันเพื่อลบข้อมูล
    # delete_all_data()